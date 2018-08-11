# -*- coding: utf-8 -*-
########################################################################################################################
#  Copyright (c) 2015 - Marcos Organizador de Negocios SRL. (<https://marcos.do/>)
#  Based in DGII Reports of Eneldo Serrata (eneldo@marcos.do)
#  Copyright (c) 2018 - SoftNet Team SRL. (<https://www.softnet.do/>)
#  Write by Manuel Gonzalez (manuel@softnet.do)
#  See LICENSE file for full copyright and licensing details.
#
# Odoo Proprietary License v1.0
#
# This software and associated files (the "Software") may only be used
# (nobody can redistribute (or sell) your module once they have bought it, unless you gave them your consent)
# if you have purchased a valid license
# from the authors, typically via Odoo Apps, or if you have received a written
# agreement from the authors of the Software (see the COPYRIGHT file).
#
# You may develop Odoo modules that use the Software as a library (typically
# by depending on it, importing it and using its resources), but without copying
# any source code or material from the Software. You may distribute those
# modules under the license of your choice, provided that this license is
# compatible with the terms of the Odoo Proprietary License (For example:
# LGPL, MIT, or proprietary licenses similar to this one).
#
# It is forbidden to publish, distribute, sublicense, or sell copies of the Software
# or modified copies of the Software.
#
# The above copyright notice and this permission notice must be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SOFTWARE.
########################################################################################################################

from odoo import models, fields, api, exceptions
from var_dump import var_dump
from pprint import pprint as pp

from openpyxl import load_workbook
import base64
import os

import re
import calendar

import logging

_logger = logging.getLogger(__name__)

try:
    from stdnum.do import ncf, rnc, cedula
except(ImportError, IOError) as err:
    _logger.debug(err)


class DgiiReport(models.Model):
    _name = "dgii.report"
    _inherit = ['mail.thread', 'ir.needaction_mixin']
    _order = "name"

    @api.multi
    @api.depends("purchase_report")
    def _purchase_report_totals(self):

        # Tipos de Bienes y Servicios Comprados
        # Columna 3 del 606
        summary_dict = {
            "01": {"count": 0, "amount": 0.0},
            "02": {"count": 0, "amount": 0.0},
            "03": {"count": 0, "amount": 0.0},
            "04": {"count": 0, "amount": 0.0},
            "05": {"count": 0, "amount": 0.0},
            "06": {"count": 0, "amount": 0.0},
            "07": {"count": 0, "amount": 0.0},
            "08": {"count": 0, "amount": 0.0},
            "09": {"count": 0, "amount": 0.0},
            "10": {"count": 0, "amount": 0.0},
            "11": {"count": 0, "amount": 0.0},
        }

        for rec in self: #self  = lines on model DgiiReportPurchaseLine
            rec.ITBIS_TOTAL = 0
            rec.ITBIS_TOTAL_NC = 0
            rec.ITBIS_TOTAL_PAYMENT = 0
            rec.TOTAL_MONTO_FACTURADO = 0
            rec.MONTO_FACTURADO_SERVICIOS = 0
            rec.MONTO_FACTURADO_BIENES = 0
            rec.TOTAL_MONTO_NC = 0
            rec.TOTAL_MONTO_PAYMENT = 0
            rec.ITBIS_RETENIDO = 0
            rec.RETENCION_RENTA = 0
            rec.ITBIS_FACTURADO_BIENES = 0
            rec.ITBIS_FACTURADO_SERVICIOS = 0
            rec.ITBIS_SUJETO_PROPORCIONALIDAD = 0
            rec.ITBIS_LLEVADO_ALCOSTO = 0
            rec.ITBIS_POR_ADELANTAR = 0
            rec.ITBIS_PERCIBIDO_COMPRAS = 0
            rec.RETENCION_RENTA = 0
            rec.ISR_PERCIBIDO_COMPRAS = 0
            rec.IMPUESTO_ISC = 0
            rec.IMPUESTO_OTROS = 0
            rec.MONTO_PROPINA_LEGAL = 0

            for purchase in rec.purchase_report:

                TIPO_COMPROBANTE = self.getTipoComprobante(purchase)

                if TIPO_COMPROBANTE == "04": # 04 = NOTAS DE CRÉDITOS #TODO check to validate NC for Monto Facturado Bienes/Servicios
                    rec.ITBIS_TOTAL_NC += purchase.ITBIS_FACTURADO_TOTAL
                    rec.TOTAL_MONTO_NC += purchase.MONTO_FACTURADO
                    rec.ITBIS_RETENIDO -= purchase.ITBIS_RETENIDO
                    rec.ITBIS_FACTURADO_SERVICIOS -= purchase.ITBIS_FACTURADO_SERVICIOS
                    rec.ITBIS_FACTURADO_BIENES -= purchase.ITBIS_FACTURADO_BIENES #TODO validate if this line is necessary
                    rec.ITBIS_SUJETO_PROPORCIONALIDAD -= purchase.ITBIS_SUJETO_PROPORCIONALIDAD #TODO validate if this line is necessary
                    rec.ITBIS_LLEVADO_ALCOSTO -= purchase.ITBIS_LLEVADO_ALCOSTO #TODO validate if this line is necessary
                    rec.ITBIS_POR_ADELANTAR -= purchase.ITBIS_POR_ADELANTAR #TODO validate if this line is necessary
                    rec.ITBIS_PERCIBIDO_COMPRAS -= purchase.ITBIS_PERCIBIDO_COMPRAS #TODO validate if this line is necessary
                    rec.RETENCION_RENTA -= purchase.RETENCION_RENTA
                    rec.ISR_PERCIBIDO_COMPRAS -= purchase.ISR_PERCIBIDO_COMPRAS #TODO validate if this line is necessary
                    rec.IMPUESTO_ISC -= purchase.IMPUESTO_ISC #TODO validate if this line is necessary
                    rec.IMPUESTO_OTROS -= purchase.IMPUESTO_OTROS #TODO validate if this line is necessary
                    rec.MONTO_PROPINA_LEGAL -= purchase.MONTO_PROPINA_LEGAL #TODO validate if this line is necessary
                elif purchase.NUMERO_COMPROBANTE_MODIFICADO == False:
                    rec.TOTAL_MONTO_FACTURADO += purchase.MONTO_FACTURADO
                    rec.MONTO_FACTURADO_SERVICIOS += purchase.MONTO_FACTURADO_SERVICIOS
                    rec.MONTO_FACTURADO_BIENES += purchase.MONTO_FACTURADO_BIENES
                    rec.ITBIS_TOTAL += purchase.ITBIS_FACTURADO_TOTAL
                    rec.ITBIS_FACTURADO_SERVICIOS += purchase.ITBIS_FACTURADO_SERVICIOS
                    rec.ITBIS_FACTURADO_BIENES += purchase.ITBIS_FACTURADO_BIENES
                    rec.ITBIS_RETENIDO += purchase.ITBIS_RETENIDO
                    rec.ITBIS_SUJETO_PROPORCIONALIDAD += purchase.ITBIS_SUJETO_PROPORCIONALIDAD
                    rec.ITBIS_LLEVADO_ALCOSTO += purchase.ITBIS_LLEVADO_ALCOSTO
                    rec.ITBIS_POR_ADELANTAR += purchase.ITBIS_POR_ADELANTAR
                    rec.ITBIS_PERCIBIDO_COMPRAS += purchase.ITBIS_PERCIBIDO_COMPRAS
                    rec.RETENCION_RENTA += purchase.RETENCION_RENTA
                    rec.ISR_PERCIBIDO_COMPRAS += purchase.ISR_PERCIBIDO_COMPRAS
                    rec.IMPUESTO_ISC += purchase.IMPUESTO_ISC
                    rec.IMPUESTO_OTROS += purchase.IMPUESTO_OTROS
                    rec.MONTO_PROPINA_LEGAL += purchase.MONTO_PROPINA_LEGAL

                summary_dict[purchase.invoice_id.expense_type]["count"] += 1
                summary_dict[purchase.invoice_id.expense_type]["amount"] += purchase.MONTO_FACTURADO

            rec.ITBIS_TOTAL_PAYMENT = rec.ITBIS_TOTAL - rec.ITBIS_TOTAL_NC
            rec.TOTAL_MONTO_PAYMENT = rec.TOTAL_MONTO_FACTURADO - rec.TOTAL_MONTO_NC
            rec.ITBIS_POR_ADELANTAR = rec.ITBIS_TOTAL - rec.ITBIS_LLEVADO_ALCOSTO

            rec.pcount_01 = summary_dict["01"]["count"]
            rec.pcount_02 = summary_dict["02"]["count"]
            rec.pcount_03 = summary_dict["03"]["count"]
            rec.pcount_04 = summary_dict["04"]["count"]
            rec.pcount_05 = summary_dict["05"]["count"]
            rec.pcount_06 = summary_dict["06"]["count"]
            rec.pcount_07 = summary_dict["07"]["count"]
            rec.pcount_08 = summary_dict["08"]["count"]
            rec.pcount_09 = summary_dict["09"]["count"]
            rec.pcount_10 = summary_dict["10"]["count"]
            rec.pcount_11 = summary_dict["11"]["count"]

            rec.pamount_01 = summary_dict["01"]["amount"]
            rec.pamount_02 = summary_dict["02"]["amount"]
            rec.pamount_03 = summary_dict["03"]["amount"]
            rec.pamount_04 = summary_dict["04"]["amount"]
            rec.pamount_05 = summary_dict["05"]["amount"]
            rec.pamount_06 = summary_dict["06"]["amount"]
            rec.pamount_07 = summary_dict["07"]["amount"]
            rec.pamount_08 = summary_dict["08"]["amount"]
            rec.pamount_09 = summary_dict["09"]["amount"]
            rec.pamount_10 = summary_dict["10"]["amount"]
            rec.pamount_11 = summary_dict["11"]["amount"]

    @api.multi
    @api.depends("sale_report")
    def _sale_report_totals(self):

        # Tipos de NCFs by name
        summary_dict = {
            "final": {"count": 0, "amount": 0.0},
            "fiscal": {"count": 0, "amount": 0.0},
            "gov": {"count": 0, "amount": 0.0},
            "special": {"count": 0, "amount": 0.0},
            "unico": {"count": 0, "amount": 0.0},
        }

        for rec in self:
            rec.SALE_ITBIS_TOTAL = 0
            rec.SALE_ITBIS_NC = 0
            rec.SALE_ITBIS_CHARGED = 0
            rec.SALE_TOTAL_MONTO_FACTURADO = 0
            rec.SALE_TOTAL_MONTO_NC = 0
            rec.SALE_TOTAL_MONTO_CHARGED = 0
            rec.MONTO_FACTURADO_EXCENTO = 0

            for sale in rec.sale_report:
                if sale.NUMERO_COMPROBANTE_FISCAL[9:-8] == "04":
                    rec.SALE_ITBIS_NC += sale.ITBIS_FACTURADO
                    rec.SALE_TOTAL_MONTO_NC += sale.MONTO_FACTURADO
                    #TODO falta manejar las notas de credito que afectan facturas de otro periodo.
                    rec.MONTO_FACTURADO_EXCENTO -= sale.MONTO_FACTURADO_EXCENTO
                else:
                    rec.SALE_ITBIS_TOTAL += sale.ITBIS_FACTURADO
                    rec.SALE_TOTAL_MONTO_FACTURADO += sale.MONTO_FACTURADO
                    rec.MONTO_FACTURADO_EXCENTO += sale.MONTO_FACTURADO_EXCENTO

                summary_dict[sale.invoice_id.sale_fiscal_type]["count"] += 1
                summary_dict[sale.invoice_id.sale_fiscal_type]["amount"] += sale.MONTO_FACTURADO

            rec.SALE_ITBIS_CHARGED = rec.SALE_ITBIS_TOTAL - rec.SALE_ITBIS_NC
            rec.SALE_TOTAL_MONTO_CHARGED = rec.SALE_TOTAL_MONTO_FACTURADO - rec.SALE_TOTAL_MONTO_NC

            rec.count_final = summary_dict["final"]["count"]

            rec.count_fiscal = summary_dict["fiscal"]["count"]
            rec.count_gov = summary_dict["gov"]["count"]
            rec.count_special = summary_dict["special"]["count"]
            rec.count_unico = summary_dict["unico"]["count"]
            rec.amount_final = summary_dict["final"]["amount"]
            rec.amount_fiscal = summary_dict["fiscal"]["amount"]
            rec.amount_gov = summary_dict["gov"]["amount"]
            rec.amount_special = summary_dict["special"]["amount"]
            rec.amount_unico = summary_dict["unico"]["amount"]

    @api.multi
    @api.depends("purchase_report", "sale_report")
    def _count_records(self):
        for rec in self:
            rec.COMPRAS_CANTIDAD_REGISTRO = rec.purchase_report and len(rec.purchase_report)
            rec.VENTAS_CANTIDAD_REGISTRO = rec.sale_report and len(rec.sale_report)
            rec.CANCEL_CANTIDAD_REGISTRO = rec.cancel_report and len(rec.cancel_report)
            rec.EXTERIOR_CANTIDAD_REGISTRO = rec.exterior_filename and len(rec.exterior_report)

    company_id = fields.Many2one('res.company', 'EMPRESA', required=False,
                                 default=lambda self: self.env.user.company_id)
    name = fields.Char(string=u"PERÍODO MES/AÑO", required=True, unique=True, index=True)
    positive_balance = fields.Float(u"SALDO A FAVOR ANTERIOR", required=True)

    it_filename = fields.Char()
    it_binary = fields.Binary(string=u"Archivo excel IT-1")

    ir17_filename = fields.Char()
    ir17_binary = fields.Binary(string=u"Archivo excel IR-17")

    # 606
    COMPRAS_CANTIDAD_REGISTRO = fields.Integer(u"Cantidad de registros", compute=_count_records)

    TOTAL_MONTO_FACTURADO = fields.Float(u"Monto compra", compute=_purchase_report_totals)
    MONTO_FACTURADO_SERVICIOS = fields.Float(u"Monto Facturado Servicios", compute=_purchase_report_totals)
    MONTO_FACTURADO_BIENES = fields.Float(u"Monto Facturado Bienes", compute=_purchase_report_totals)

    ITBIS_TOTAL = fields.Float(u"ITBIS Compras", compute=_purchase_report_totals)
    ITBIS_FACTURADO_SERVICIOS = fields.Float(u"ITBIS Facturado Servicios", compute=_purchase_report_totals)
    ITBIS_FACTURADO_BIENES = fields.Float(u"ITBIS Facturado Bienes", compute=_purchase_report_totals)

    TOTAL_MONTO_NC = fields.Float(u"Notas de crédito", compute=_purchase_report_totals)
    ITBIS_TOTAL_NC = fields.Float(u"ITBIS Notas de crédito", compute=_purchase_report_totals)

    TOTAL_MONTO_PAYMENT = fields.Float(u"Total monto facturado", compute=_purchase_report_totals)
    ITBIS_TOTAL_PAYMENT = fields.Float(u"ITBIS Pagado", compute=_purchase_report_totals)

    ITBIS_RETENIDO = fields.Float(u"ITBIS Retenido", compute=_purchase_report_totals)
    RETENCION_RENTA = fields.Float(u"Retención Renta", compute=_purchase_report_totals)

    purchase_report = fields.One2many(u"dgii.report.purchase.line", "dgii_report_id")
    purchase_filename = fields.Char()
    purchase_binary = fields.Binary(string=u"Archivo 606 TXT")

    # 606 type summary
    currency_id = fields.Many2one(related="company_id.currency_id")

    pcount_01 = fields.Integer(compute=_purchase_report_totals)
    pcount_02 = fields.Integer(compute=_purchase_report_totals)
    pcount_03 = fields.Integer(compute=_purchase_report_totals)
    pcount_04 = fields.Integer(compute=_purchase_report_totals)
    pcount_05 = fields.Integer(compute=_purchase_report_totals)
    pcount_06 = fields.Integer(compute=_purchase_report_totals)
    pcount_07 = fields.Integer(compute=_purchase_report_totals)
    pcount_08 = fields.Integer(compute=_purchase_report_totals)
    pcount_09 = fields.Integer(compute=_purchase_report_totals)
    pcount_10 = fields.Integer(compute=_purchase_report_totals)
    pcount_11 = fields.Integer(compute=_purchase_report_totals)

    pamount_01 = fields.Monetary(compute=_purchase_report_totals)
    pamount_02 = fields.Monetary(compute=_purchase_report_totals)
    pamount_03 = fields.Monetary(compute=_purchase_report_totals)
    pamount_04 = fields.Monetary(compute=_purchase_report_totals)
    pamount_05 = fields.Monetary(compute=_purchase_report_totals)
    pamount_06 = fields.Monetary(compute=_purchase_report_totals)
    pamount_07 = fields.Monetary(compute=_purchase_report_totals)
    pamount_08 = fields.Monetary(compute=_purchase_report_totals)
    pamount_09 = fields.Monetary(compute=_purchase_report_totals)
    pamount_10 = fields.Monetary(compute=_purchase_report_totals)
    pamount_11 = fields.Monetary(compute=_purchase_report_totals)

    # 607
    VENTAS_CANTIDAD_REGISTRO = fields.Integer(u"Cantidad de registros", compute=_count_records)

    SALE_TOTAL_MONTO_FACTURADO = fields.Float(u"Total Facturado", compute=_sale_report_totals)
    SALE_ITBIS_TOTAL = fields.Float(u"ITBIS ventas", compute=_sale_report_totals)

    SALE_TOTAL_MONTO_NC = fields.Float(u"Total Notas de crédito", compute=_sale_report_totals)
    SALE_ITBIS_NC = fields.Float(u"ITBIS Notas de crédito", compute=_sale_report_totals)

    SALE_TOTAL_MONTO_CHARGED = fields.Float(u"Facturado", compute=_sale_report_totals)
    SALE_ITBIS_CHARGED = fields.Float(u"ITBIS Cobrado", compute=_sale_report_totals)
    MONTO_FACTURADO_EXCENTO = fields.Float(u"ITBIS Cobrado", compute=_sale_report_totals)

    sale_filename = fields.Char()
    sale_binary = fields.Binary(string=u"Archivo 607 TXT")

    sale_report = fields.One2many("dgii.report.sale.line", "dgii_report_id")

    # 607 type summary
    count_final = fields.Integer(compute=_sale_report_totals)
    count_fiscal = fields.Integer(compute=_sale_report_totals)
    count_gov = fields.Integer(compute=_sale_report_totals)
    count_special = fields.Integer(compute=_sale_report_totals)
    count_unico = fields.Integer(compute=_sale_report_totals)
    amount_final = fields.Integer(compute=_sale_report_totals)
    amount_fiscal = fields.Integer(compute=_sale_report_totals)
    amount_gov = fields.Integer(compute=_sale_report_totals)
    amount_special = fields.Integer(compute=_sale_report_totals)
    amount_unico = fields.Integer(compute=_sale_report_totals)

    # 608
    CANCEL_CANTIDAD_REGISTRO = fields.Integer(u"Cantidad de registros", compute=_count_records)
    cancel_report = fields.One2many("dgii.cancel.report.line", "dgii_report_id")
    cancel_filename = fields.Char()
    cancel_binary = fields.Binary(string=u"Archivo 608 TXT")

    # 609
    EXTERIOR_CANTIDAD_REGISTRO = fields.Integer(u"Cantidad de registros", compute=_count_records)
    EXTERIOR_TOTAL_MONTO_FACTURADO = fields.Float()
    exterior_report = fields.One2many("dgii.exterior.report.line", "dgii_report_id")
    exterior_filename = fields.Char(u"Total Monto Facturado")
    exterior_binary = fields.Binary(string=u"Archivo 607 TXT")

    state = fields.Selection([('draft', 'Nuevo'), ('error', 'Con errores'), ('done', 'Validado')], default="draft")

    def get_invoice_in_draft_error(self, invoice_ids):
        error_list = {}
        error_msg = "Factura sin validar"
        for invoice_id in invoice_ids:
            if not error_list.get(invoice_id.id, False):
                error_list.update(
                    {invoice_id.id: [
                        (invoice_id.type, invoice_id.number, error_msg)]})
            else:
                error_list[invoice_id.id].append(
                    (invoice_id.type, invoice_id.number, error_msg))
        return error_list


    ''''
        With this method they want get all invoices paid in a period of time
        and use them in the report of the current month (period: start and end date given).
        But, acording with some accountants, this should be only valid for invoices
        with retention of ITBIS and ISR and of kind "Informal", which means that
        don't matter if the NCF is issued by the provider or by the company requiring
        the services, what matter is the document/identification of the provider,
        if this is of kind of "cedula", so it is informal.
    '''
    def get_late_informal_payed_invoice(self, start_date, end_date):

        invoice_ids = self.env["account.invoice"] # this is like define an empty array|object

        paid_invoice_ids = self.env["account.payment"].search(
            [('payment_date', '>=', start_date), ('payment_date', '<=', end_date), ('invoice_ids', '!=', False)])

        for paid_invoice_id in paid_invoice_ids:
            RNC_CEDULA, TIPO_IDENTIFICACION = self.get_identification_info(paid_invoice_id.partner_id.vat)
            if TIPO_IDENTIFICACION == "2": # just informal with or without ncf given.
                account_move_lines = self.env["account.move.line"].search([('payment_id', '=', paid_invoice_id.id)])
                if(account_move_lines):
                    invoice = account_move_lines[0].invoice_id
                    FECHA_PAGO, ITBIS_RETENIDO, RETENCION_RENTA, TIPO_RETENCION_ISR = self.get_payment_date_and_retention_data(invoice)
                    if ITBIS_RETENIDO or RETENCION_RENTA:
                        invoice_ids |= paid_invoice_id.invoice_ids.filtered(lambda r: r.journal_id.purchase_type in ("informal", "normal")).filtered(lambda r: r.journal_id.type == "purchase") # this is like array_push(), just making appends

        return invoice_ids

    def get_identification_info(self, vat):
        RNC_CEDULA = vat and re.sub("[^0-9]", "", vat.strip()) or False
        TIPO_IDENTIFICACION = "3"

        if RNC_CEDULA:
            if len(RNC_CEDULA) == 9:
                TIPO_IDENTIFICACION = "1"
            elif len(RNC_CEDULA) == 11:
                TIPO_IDENTIFICACION = "2"

        if TIPO_IDENTIFICACION == "3":
            RNC_CEDULA = ""

        return RNC_CEDULA, TIPO_IDENTIFICACION

    # def validate_fiscal_information(self, vat, ncf, invoice_type, origin_invoice_ids): # AttributeError: 'account.invoice' object has no attribute 'origin_invoice_ids', even in the old marcos addons

    def validate_fiscal_information(self, vat, invoice):

        error_list = []

        if invoice.type == 'out_invoice':
            vat = invoice.company_id.vat

        if vat and len(vat) == 9 and not rnc.is_valid(vat):
            error_list.append(u"El RNC no es válido")

        if vat and len(vat) == 11 and not cedula.is_valid(vat):
            error_list.append(u"La Cédula no es válida")

        if not ncf.is_valid(invoice.number) or not ncf.check_dgii(vat, invoice.number):
            error_list.append(u"El NCF no es válido.  RNC: %s y tipo de Factura: %s" % (vat, invoice.type))

        # if len(origin_invoice_ids) > 1 and invoice_type in ("out_refund", "in_refund"):
        #     error_list.append(u"NC/ND Afectando varias facturas")

        # if not origin_invoice_ids and invoice_type in ("out_refund", "in_refund"):
        #     error_list.append(u"NC/ND sin comprobante que afecta")

        if invoice.type in ("out_refund", "in_refund"):
            error_list.append(u"NC/ND sin comprobante que afecta")

        if not invoice.number:
            error_list.append(u"Factura validada sin número asignado")

        if invoice.type == 'in_invoice' and not invoice.expense_type:
            error_list.append(u"La factura %s no tiene especificado el tipo de costos y gastos requerído por el DGII." % invoice.number)

        return error_list

    # 608
    @api.multi
    def create_cancel_invoice_lines(self, cancel_invoice_ids):
        self.cancel_report.unlink()
        new_cancel_report = []
        cancel_line = 1
        for invoice_id in cancel_invoice_ids:
            new_cancel_report.append((0, 0, {"LINE": cancel_line, "TIPO_ANULACION": invoice_id.anulation_type,
                                             "FECHA_COMPROBANTE": invoice_id.date_invoice,
                                             "NUMERO_COMPROBANTE_FISCAL": invoice_id.move_name}))
            self.write({"cancel_report": new_cancel_report})
            cancel_line += 1

    @api.multi
    def get_numero_de_comprobante_modificado(self, invoice_id):
        NUMERO_COMPROBANTE_MODIFICADO = False
        AFFECTED_NVOICE_ID = False

        origin_invoice_id = invoice_id.origin_invoice_ids.filtered(lambda x: x.state in ("open", "paid"))

        if not origin_invoice_id:
            origin_invoice_id = self.env["account.invoice"].search(
                [('number', '=', invoice_id.origin)])

        NUMERO_COMPROBANTE_MODIFICADO = origin_invoice_id[0].number
        AFFECTED_NVOICE_ID = origin_invoice_id[0].id

        return NUMERO_COMPROBANTE_MODIFICADO, AFFECTED_NVOICE_ID


    '''
    *** This method is only called when the Invoice is paid
    '''
    def get_payment_date_and_retention_data(self, invoice_id):

        FECHA_PAGO = False
        ITBIS_RETENIDO = 0
        RETENCION_RENTA = 0
        TIPO_RETENCION_ISR = False

        if invoice_id.id == False : #TODO for some reason, invoice_id has not any properties some times...
            return FECHA_PAGO, ITBIS_RETENIDO, RETENCION_RENTA, TIPO_RETENCION_ISR

        # payment_rel = self.env["account.invoice.payment.rel"].search(['invoice_id', '=', invoice_id.id]) # This return an error:  KeyError: 'account.invoice.payment.rel'
        self.env.cr.execute("select * from account_invoice_payment_rel where invoice_id = %s" % invoice_id.id)
        payment_rel = self.env.cr.dictfetchone() # return just one diccionario, like laravel: ->first()

        payment = self.env["account.payment"].browse(payment_rel['payment_id'])
        FECHA_PAGO = payment.payment_date

        move_id = self.env["account.move.line"].search([("move_id", "=", invoice_id.move_id.id), ('full_reconcile_id', '!=', False)]) # just one is full_reconcile_id

        if invoice_id.journal_id.purchase_type in ("informal", "normal"):

            if move_id:

                ''' I commented the below query because when I run in my DB:
                select * from account_move_line where payment_id > 0 and invoice_id > 0 order by payment_id desc;
                I just get four invoice and I have other Invoice with ITBIS and ISR retentions but it doesn't appears in this search....
                '''
                # account_move_lines = self.env["account.move.line"].search(
                #     [('invoice_id', '=', invoice_id.id), ('payment_id', '!=', False),
                #      ('tax_line_id', '!=', False)])

                account_move_lines = self.env["account.move.line"].search(
                    [('move_id', '=', invoice_id.move_id.id),('tax_line_id', '!=', False)]) # I removed the filter ('payment_id', '!=', False) because in one of my case the move lines don't have payment_id, why?, I don't have idea....

                if account_move_lines:
                    for line in account_move_lines:
                        if line.tax_line_id.purchase_tax_type == "ritbis":
                            ITBIS_RETENIDO += line.credit
                        elif line.tax_line_id.purchase_tax_type == "isr":
                            RETENCION_RENTA += line.credit
                            TIPO_RETENCION_ISR = line.tax_line_id.isr_retention_type or False

        return FECHA_PAGO, ITBIS_RETENIDO, RETENCION_RENTA, TIPO_RETENCION_ISR

    '''
        This method return:
        Impuesto Selectivo al Consumo (Casilla 20),
        Otros Impuesto/Tasas (Casilla 21),
        and Monto Propina Legal (Casilla 22)
        but only when the invoice is Open or Paid
    '''
    def get_isc_propina_otros(self, invoice_id):

        IMPUESTO_ISC = 0
        IMPUESTO_OTROS = 0
        MONTO_PROPINA_LEGAL = 0

        if invoice_id.id == False : #TODO for some reason, invoice_id has not any properties some times...
            return IMPUESTO_ISC, IMPUESTO_OTROS, MONTO_PROPINA_LEGAL

        if invoice_id.state in ("open", "paid"):

            account_move_lines = self.env["account.move.line"].search(
                [('move_id', '=', invoice_id.move_id.id),('tax_line_id', '!=', False)])

            if account_move_lines:
                for line in account_move_lines:
                    if line.tax_line_id.purchase_tax_type == "isc":
                        IMPUESTO_ISC += line.credit
                    elif line.tax_line_id.purchase_tax_type in ("cdt"): #TODO may be there another taxes as "IMPUESTOS_OTROS" that are not just CDT.
                        IMPUESTO_OTROS += line.credit
                    elif line.tax_line_id.purchase_tax_type in ("propina_legal"):
                        MONTO_PROPINA_LEGAL += line.credit

        return IMPUESTO_ISC, IMPUESTO_OTROS, MONTO_PROPINA_LEGAL


    def get_format_pago(self, invoice_id):

        FORMA_PAGO = '04' # 04 = COMPRA A CREDITO

        if invoice_id.state == "paid":
            
            self.env.cr.execute("select * from account_invoice_payment_rel where invoice_id = %s" % invoice_id.id)
            payment_rel = self.env.cr.dictfetchall() # return an array of dicts, like laravel: ->get()

            if len(payment_rel) == 0: # could be a NOTA DE CREDITO, they don't seems store payment_id

                refund_invoice_id = self.env["account.invoice"].search([('refund_invoice_id', '=', invoice_id.id)])
                if refund_invoice_id:
                    FORMA_PAGO = '04'
                
            elif len(payment_rel) > 1:
        
                FORMA_PAGO = '07' # 07 = MIXTO

            else:

                payment = self.env["account.payment"].browse(payment_rel[0]['payment_id'])

                if payment.writeoff_account_id > 0: #TODO validate with an accountant this.
                    FORMA_PAGO = '07' # 07 = MIXTO
                elif payment.journal_id.payment_form == 'cash':
                    FORMA_PAGO = '01'
                elif payment.journal_id.payment_form == 'bank':
                    FORMA_PAGO = '02'
                elif payment.journal_id.payment_form == 'card':
                    FORMA_PAGO = '03'
                elif payment.journal_id.payment_form == 'credit': # just in case they have a journal of credit
                    FORMA_PAGO = '04'
                elif payment.journal_id.payment_form == 'swap':
                    FORMA_PAGO = '05' # Permuta

        return FORMA_PAGO


    @api.multi
    def create_sales_lines(self, data):
        dataText = ','.join(self.env.cr.mogrify('(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)', row) for row in data)

        sale_insert_sql = """
                        INSERT INTO dgii_report_sale_line ("dgii_report_id","LINE","RNC_CEDULA","TIPO_IDENTIFICACION",
                        "NUMERO_COMPROBANTE_FISCAL","NUMERO_COMPROBANTE_MODIFICADO","FECHA_COMPROBANTE","ITBIS_FACTURADO","MONTO_FACTURADO",
                        "MONTO_FACTURADO_EXCENTO","invoice_id","affected_nvoice_id","nc") values {}
                        """.format(dataText)
        self.env.cr.execute(sale_insert_sql)

    @api.multi
    def create_purchase_lines(self, data):
        dataText = ','.join(
            self.env.cr.mogrify('(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)', row) for row in data)

        purchase_insert_sql = """
                            INSERT INTO dgii_report_purchase_line ("dgii_report_id",
                            "LINE",
                            "RNC_CEDULA",
                            "TIPO_IDENTIFICACION",
                            "NUMERO_COMPROBANTE_FISCAL",
                            "NUMERO_COMPROBANTE_MODIFICADO",
                            "FECHA_COMPROBANTE",
                            "FECHA_PAGO",
                            "TIPO_BIENES_SERVICIOS_COMPRADOS",
                            "MONTO_FACTURADO",
                            "MONTO_FACTURADO_SERVICIOS",
                            "MONTO_FACTURADO_BIENES",
                            "ITBIS_FACTURADO_TOTAL",
                            "ITBIS_FACTURADO_BIENES",
                            "ITBIS_FACTURADO_SERVICIOS",
                            "ITBIS_RETENIDO",
                            "ITBIS_SUJETO_PROPORCIONALIDAD",
                            "ITBIS_LLEVADO_ALCOSTO",
                            "ITBIS_POR_ADELANTAR",
                            "ITBIS_PERCIBIDO_COMPRAS",
                            "TIPO_RETENCION_ISR",
                            "RETENCION_RENTA",
                            "ISR_PERCIBIDO_COMPRAS",
                            "IMPUESTO_ISC",
                            "IMPUESTO_OTROS",
                            "MONTO_PROPINA_LEGAL",
                            "FORMA_PAGO",
                            "invoice_id",
                            "affected_nvoice_id",
                            "nc") values {}
                            """.format(dataText)
        self.env.cr.execute(purchase_insert_sql)

    @api.multi
    def post_error_list(self, error_list):
        out_inovice_url = "/web#id={}&view_type=form&model=account.invoice&action=196"
        in_inovice_url = "/web#id={}&view_type=form&model=account.invoice&menu_id=119&action=197"
        if error_list:
            message = "<ul>"
            for ncf, errors in error_list.iteritems():
                message += "<li>{}</li><ul>".format(errors[0][1] or "Factura invalida")
                for error in errors:
                    if error[0] in ("out_invoice", "out_refund"):
                        message += u"<li><a target='_blank' href='{}'>{}</a></li>".format(out_inovice_url.format(ncf),
                                                                                          error[2])
                    else:
                        message += u"<li><a target='_blank' href='{}'>{}</a></li>".format(in_inovice_url.format(ncf),
                                                                                          error[2])
                message += "</ul>"
            message += "</ul>"

            self.message_post(body=message)
            self.state = "error"
        else:
            self.message_post(body="Generado correctamente")
            self.state = "done"

    @api.multi
    def generate_report(self):

        try:
            month, year = self.name.split("/")
            last_day = calendar.monthrange(int(year), int(month))[1]
            start_date = "{}-{}-01".format(year, month)
            end_date = "{}-{}-{}".format(year, month, last_day)
        except:
            raise exceptions.ValidationError(u"Período inválido")

        self.purchase_report.unlink()
        self.sale_report.unlink()
        self.cancel_report.unlink()
        self.exterior_report.unlink()

        self.it_filename = False
        self.it_binary = False
        self.ir17_filename = False
        self.ir17_binary = False

        self.sale_binary = False
        self.sale_filename = False
        self.purchase_binary = False
        self.purchase_filename = False
        self.cancel_binary = False
        self.cancel_filename = False

        xls_dict = {"it1": {}, "ir17": {}}
        purchase_report = []
        sale_report = []
        ext_report = []
        sale_line = 1
        purchase_line = 1
        ext_line = 1

        sale_except_tax_id = self.env.ref("l10n_do.{}_tax_0_sale".format(self.company_id.id))
        purchase_except_tax_id = self.env.ref("l10n_do.{}_tax_0_purch".format(self.company_id.id))
        untax_ids = (sale_except_tax_id.id, purchase_except_tax_id.id)

        journal_ids = self.env["account.journal"].search(
            ['|', ('ncf_control', '=', True), ('ncf_remote_validation', '=', True)])

        # searching invoices to this period
        invoice_ids = self.env["account.invoice"].search(
            [('date_invoice', '>=', start_date), ('date_invoice', '<=', end_date),
             ('journal_id', 'in', journal_ids.ids)])

        error_list = self.get_invoice_in_draft_error(invoice_ids.filtered(lambda x: x.state == "draft"))

        self.create_cancel_invoice_lines(invoice_ids.filtered(lambda x: x.state == 'cancel' and
                                                                        x.type in ("out_invoice", "out_refund") and
                                                                        x.move_name))

        invoice_ids = invoice_ids.filtered(lambda x: x.state in ('open', 'paid'))

        invoice_ids |= self.get_late_informal_payed_invoice(start_date, end_date)

        count = len(invoice_ids)
        for invoice_id in invoice_ids:

            RNC_CEDULA, TIPO_IDENTIFICACION = self.get_identification_info(invoice_id.partner_id.vat)

            # error_msg = self.validate_fiscal_information(RNC_CEDULA, invoice_id.number, invoice_id.type,
                                                        #  invoice_id.origin_invoice_ids) # AttributeError: 'account.invoice' object has no attribute 'origin_invoice_ids'.  Even in the old marcos addons

            error_msg = self.validate_fiscal_information(RNC_CEDULA, invoice_id)

            if error_msg:
                for error in error_msg:
                    if not error_list.get(invoice_id.id, False):
                        error_list.update({invoice_id.id: [(invoice_id.type, invoice_id.number, error)]})
                    else:
                        error_list[invoice_id.id].append((invoice_id.type, invoice_id.number, error))
                continue

            NUMERO_COMPROBANTE_FISCAL = invoice_id.number
            FECHA_COMPROBANTE = invoice_id.date_invoice

            NUMERO_COMPROBANTE_MODIFICADO = AFFECTED_NVOICE_ID = False

            IMPUESTO_ISC, IMPUESTO_OTROS, MONTO_PROPINA_LEGAL = self.get_isc_propina_otros(invoice_id)

            if invoice_id.type in ("out_refund", "in_refund"):
                NUMERO_COMPROBANTE_MODIFICADO, AFFECTED_NVOICE_ID = self.get_numero_de_comprobante_modificado(invoice_id)

            FECHA_PAGO = ITBIS_RETENIDO = RETENCION_RENTA = TIPO_RETENCION_ISR = False

            if invoice_id.state == "paid":
                FECHA_PAGO, ITBIS_RETENIDO, RETENCION_RENTA, TIPO_RETENCION_ISR = self.get_payment_date_and_retention_data(invoice_id)
                invoiceMonth = int(invoice_id.date_invoice[5:7])
                paidMonth = int(FECHA_PAGO[5:7]) if FECHA_PAGO else False
                periodMonth = int(month)

                '''
                    With the validation below we are looking don't show payment date or retentions info in a period
                    that the invoice was not paid yet.
                '''
                if invoiceMonth != paidMonth and invoiceMonth == periodMonth:
                    FECHA_PAGO = ITBIS_RETENIDO = RETENCION_RENTA = False

            FORMA_PAGO = self.get_format_pago(invoice_id) if invoice_id.type in ("in_invoice", "in_refund") else False

            # ''' This is one line in 606 or 607 report '''
            commun_data = {
                "RNC_CEDULA": RNC_CEDULA,
                "TIPO_IDENTIFICACION": TIPO_IDENTIFICACION,
                "TIPO_BIENES_SERVICIOS_COMPRADOS": invoice_id.expense_type,
                "NUMERO_COMPROBANTE_FISCAL": NUMERO_COMPROBANTE_FISCAL,
                "NUMERO_COMPROBANTE_MODIFICADO": NUMERO_COMPROBANTE_MODIFICADO,
                "FECHA_COMPROBANTE": FECHA_COMPROBANTE,
                "FECHA_PAGO": FECHA_PAGO and FECHA_PAGO or None,
                "invoice_id": invoice_id.id,
                "inv_partner": invoice_id.partner_id.id,
                "affected_nvoice_id": AFFECTED_NVOICE_ID,
                "nc": True if AFFECTED_NVOICE_ID else False,
                "MONTO_FACTURADO_EXCENTO": 0,
                "MONTO_FACTURADO": 0,
                "MONTO_FACTURADO_SERVICIOS": 0,
                "MONTO_FACTURADO_BIENES": 0,
                "ITBIS_FACTURADO": 0, # used in 607 report as total
                "ITBIS_FACTURADO_TOTAL": 0, # used in 606 report
                "ITBIS_FACTURADO_SERVICIOS": 0,
                "ITBIS_FACTURADO_BIENES": 0,
                "ITBIS_RETENIDO": ITBIS_RETENIDO or 0,
                "ITBIS_SUJETO_PROPORCIONALIDAD": 0,
                "ITBIS_LLEVADO_ALCOSTO": 0,
                "ITBIS_POR_ADELANTAR": 0,
                "ITBIS_PERCIBIDO_COMPRAS": 0,
                "TIPO_RETENCION_ISR": TIPO_RETENCION_ISR,
                "RETENCION_RENTA": RETENCION_RENTA or 0,
                "ISR_PERCIBIDO_COMPRAS": 0,
                "IMPUESTO_ISC": IMPUESTO_ISC,
                "IMPUESTO_OTROS": IMPUESTO_OTROS,
                "MONTO_PROPINA_LEGAL": MONTO_PROPINA_LEGAL,
                "FORMA_PAGO": FORMA_PAGO
            }

            # invoice_line_ids is the related table: account_invoice_line; this table has invoice_id column
            # invoice_line_tax_ids is the related table: account_invoice_line_tax; this table has invoice_line_id column that reference to account_invoice_line
            no_tax_line = invoice_id.invoice_line_ids.filtered(lambda x: not x.invoice_line_tax_ids)

            if no_tax_line:
                if invoice_id.type in ("out_invoice", "out_refund"):
                    no_tax_line.write({"invoice_line_tax_ids": [(4, sale_except_tax_id.id, False)]})
                else:
                    no_tax_line.write({"invoice_line_tax_ids": [(4, purchase_except_tax_id.id, False)]})

            untaxed_lines = invoice_id.invoice_line_ids.filtered(lambda x: x.invoice_line_tax_ids[0].id in untax_ids)

            untaxed_move_lines = []
            for untaxed_line in untaxed_lines:
                if invoice_id.type in ("in_invoice", 'out_refund'):
                    domain = [('move_id', '=', invoice_id.move_id.id), ('product_id', '=', untaxed_line.product_id.id),
                              ('debit', '=', abs(untaxed_line.price_subtotal_signed))]
                else:
                    domain = [('move_id', '=', invoice_id.move_id.id), ('product_id', '=', untaxed_line.product_id.id),
                              ('credit', '=', abs(untaxed_line.price_subtotal_signed))]

                move_lines = self.env["account.move.line"].search(domain)
                if move_lines:
                    untaxed_move_lines.append(move_lines)

            if untaxed_move_lines:
                if invoice_id.type in ("out_invoice", "out_refund"):
                    if not sale_except_tax_id in [t.tax_id for t in invoice_id.tax_line_ids]:
                        invoice_id.write({"tax_line_ids": [(0, 0, {"tax_id": sale_except_tax_id.id,
                                                                   "name": sale_except_tax_id.name,
                                                                   "account_id": untaxed_move_lines[
                                                                       0].account_id.id})]})
                else:
                    if not purchase_except_tax_id in [t.tax_id for t in invoice_id.tax_line_ids]:
                        invoice_id.write({"tax_line_ids": [(0, 0, {"tax_id": purchase_except_tax_id.id,
                                                                   "name": purchase_except_tax_id.name,
                                                                   "account_id": untaxed_move_lines[
                                                                       0].account_id.id})]})

                commun_data["MONTO_FACTURADO_EXCENTO"] = self.env.user.company_id.currency_id.round(
                    sum(abs(rec.debit - rec.credit) for rec in untaxed_move_lines))

            taxed_lines = invoice_id.invoice_line_ids.filtered(lambda x: x.invoice_line_tax_ids[0].id not in untax_ids)

            taxed_lines_name = [rec.product_id.id for rec in taxed_lines] # return an array of ids de products

            # _logger.warning("************* taxed_lines_name: %s" % taxed_lines_name)

            if commun_data["MONTO_FACTURADO_EXCENTO"]:
                taxed_lines_amount = self.env["account.move.line"].search(
                    [('move_id', '=', invoice_id.move_id.id), ('product_id', 'in', taxed_lines_name),
                     ("id", 'not in', [x.id for x in untaxed_move_lines])])
            else:
                taxed_lines_amount = self.env["account.move.line"].search([('move_id', '=', invoice_id.move_id.id),
                                                                           ('product_id', 'in', taxed_lines_name),
                                                                           ('tax_line_id', '=', False), #TODO, improve this filtering; with it we are looking fixing a issue in 607 report with invoice line without product selected.
                                                                           ('name', '!=', '/') #TODO, improve this filtering; with it we are looking fixing a issue in 607 report with invoice line without product selected.
                                                                    ])

            commun_data["MONTO_FACTURADO"] = self.env.user.company_id.currency_id.round(
                sum(abs(rec.debit - rec.credit) for rec in taxed_lines_amount))

            commun_data["MONTO_FACTURADO"] += commun_data["MONTO_FACTURADO_EXCENTO"]

            for tax in invoice_id.tax_line_ids: # those are ids on the table: account_invoice_tax
                tax_base_amount = commun_data["MONTO_FACTURADO"]
                untax_base_amount = commun_data["MONTO_FACTURADO_EXCENTO"]

                tax_line = self.env["account.move.line"].search(
                    [('move_id', '=', invoice_id.move_id.id), ('tax_line_id', '=', tax.tax_id.id)])

                if tax_line:
                    tax_amount = self.env.user.company_id.currency_id.round(
                        sum(abs(rec.debit - rec.credit) for rec in tax_line))

                    # if tax.tax_id.type_tax_use == "sale" or (tax.tax_id.type_tax_use == "purchase" and tax.tax_id.account_id.code == '11080101'): # 11080101 = ITBIS Pagado en Compras Locales (DEPRECATED)
                    if tax.tax_id.type_tax_use == "sale" or (tax.tax_id.type_tax_use == "purchase" and tax.tax_id.purchase_tax_type =="itbis"):
                        commun_data["ITBIS_FACTURADO"] += tax_amount # used to 607 report as total.
                        commun_data["ITBIS_FACTURADO_BIENES"] += tax_amount # used to 606 report

                    # if tax.tax_id.type_tax_use == "purchase" and tax.tax_id.account_id.code == '11080102': # 11080102 = ITBIS Pagado en Servicios Locales (DEPRECATED)
                    if tax.tax_id.type_tax_use == "purchase" and tax.tax_id.purchase_tax_type == "itbis_servicios":
                        commun_data["ITBIS_FACTURADO_SERVICIOS"] += tax_amount # used to 606 report
                else:
                    tax_amount = 0

                if invoice_id.type in ("out_refund", "in_refund"):
                    tax_base_amount = tax_base_amount * -1
                    untax_base_amount = untax_base_amount * -1
                    tax_amount = tax_amount*-1


            '''
            ********** Getting MONTO_FACTURADO_SERVICIOS and MONTO_FACTURADO_BIENES for 606 Report  **********
            '''
            if invoice_id.type in ("in_invoice", "in_refund"):

                account_move_lines = self.env["account.move.line"].search([('move_id', '=', invoice_id.move_id.id), ('product_id', '!=', False)])

                for account_move_line in account_move_lines:
                    if account_move_line.product_id.product_tmpl_id.type in ("service"):
                        commun_data["MONTO_FACTURADO_SERVICIOS"] += self.env.user.company_id.currency_id.round(abs(account_move_line.debit - account_move_line.credit))
                    else:
                        commun_data["MONTO_FACTURADO_BIENES"] += self.env.user.company_id.currency_id.round(abs(account_move_line.debit - account_move_line.credit))


                #TODO commented in new ln10 dominicana version
                # if tax.tax_id.base_it1_cels:
                #     xls_cels = tax.tax_id.base_it1_cels.split(",")

                #     for xls_cel in xls_cels:
                #         if tax.tax_id.amount == 0:
                #             if not xls_dict["it1"].get(xls_cel, False):
                #                 xls_dict["it1"].update({xls_cel: untax_base_amount})
                #             else:
                #                 xls_dict["it1"][xls_cel] += untax_base_amount
                #         else:
                #             if not xls_dict["it1"].get(xls_cel, False):
                #                 xls_dict["it1"].update({xls_cel: tax_base_amount})
                #             else:
                #                 xls_dict["it1"][xls_cel] += tax_base_amount

                #TODO commented in new ln10 dominicana version
                # if tax.tax_id.base_ir17_cels:
                #     xls_cels = tax.tax_id.base_ir17_cels.split(u",")

                #     for xls_cel in xls_cels:
                #         xls_cel = xls_cel.split(u"%")

                #         if len(xls_cel) == 1:
                #             if not xls_dict["ir17"].get(xls_cel[0], False):
                #                 xls_dict["ir17"].update({xls_cel[0]: commun_data["MONTO_FACTURADO"]})
                #             else:
                #                 xls_dict["ir17"][xls_cel[0]] += commun_data["MONTO_FACTURADO"]
                #         elif len(xls_cel) == 2:
                #             if not xls_dict["ir17"].get(xls_cel[0], False):
                #                 xls_dict["ir17"].update(
                #                     {xls_cel[0]: round(commun_data["MONTO_FACTURADO"] * (float(xls_cel[1]) / 100), 2)})
                #             else:
                #                 xls_dict["ir17"][xls_cel[0]] += round(
                #                     commun_data["MONTO_FACTURADO"] * (float(xls_cel[1]) / 100), 2)

                #TODO commented in new ln10 dominicana version
                # if tax.tax_id.tax_it1_cels:
                #     xls_cels = tax.tax_id.tax_it1_cels.split(",")
                #     for xls_cel in xls_cels:
                #         if not xls_dict["it1"].get(xls_cel, False):
                #             xls_dict["it1"].update({xls_cel: tax_amount})
                #         else:
                #             xls_dict["it1"][xls_cel] += tax_amount

                # if tax.tax_id.tax_ir17_cels:
                #     xls_cels = tax.tax_id.tax_ir17_cels.split(",")
                #     for xls_cel in xls_cels:
                #         if not xls_dict["ir17"].get(xls_cel, False):
                #             xls_dict["ir17"].update({xls_cel: tax_amount})
                #         else:
                #             xls_dict["ir17"][xls_cel] += tax_amount


            if invoice_id.type in ("out_invoice", "out_refund") and commun_data["MONTO_FACTURADO"]:
                sale_report.append((self.id,
                                    sale_line,
                                    commun_data["RNC_CEDULA"],
                                    commun_data["TIPO_IDENTIFICACION"],
                                    commun_data["NUMERO_COMPROBANTE_FISCAL"],
                                    commun_data["NUMERO_COMPROBANTE_MODIFICADO"] and commun_data[
                                        "NUMERO_COMPROBANTE_MODIFICADO"] or None,
                                    commun_data["FECHA_COMPROBANTE"],
                                    commun_data["ITBIS_FACTURADO"],
                                    commun_data["MONTO_FACTURADO"],
                                    commun_data["MONTO_FACTURADO_EXCENTO"],
                                    invoice_id.id,
                                    AFFECTED_NVOICE_ID and AFFECTED_NVOICE_ID or None,
                                    AFFECTED_NVOICE_ID and True or False))
                sale_line += 1
            elif invoice_id.type in ("in_invoice", "in_refund") and commun_data["MONTO_FACTURADO"]:

                commun_data["ITBIS_FACTURADO_TOTAL"] = commun_data["ITBIS_FACTURADO_BIENES"] + commun_data["ITBIS_FACTURADO_SERVICIOS"]

                purchase_report.append((self.id,
                                        purchase_line,
                                        commun_data["RNC_CEDULA"],
                                        commun_data["TIPO_IDENTIFICACION"],
                                        commun_data["NUMERO_COMPROBANTE_FISCAL"],
                                        commun_data["NUMERO_COMPROBANTE_MODIFICADO"] and commun_data[
                                            "NUMERO_COMPROBANTE_MODIFICADO"] or None,
                                        commun_data["FECHA_COMPROBANTE"],
                                        commun_data["FECHA_PAGO"] and commun_data["FECHA_PAGO"] or None,
                                        commun_data["TIPO_BIENES_SERVICIOS_COMPRADOS"],
                                        commun_data["MONTO_FACTURADO"],
                                        commun_data["MONTO_FACTURADO_SERVICIOS"],
                                        commun_data["MONTO_FACTURADO_BIENES"],
                                        commun_data["ITBIS_FACTURADO_TOTAL"],
                                        commun_data["ITBIS_FACTURADO_BIENES"],
                                        commun_data["ITBIS_FACTURADO_SERVICIOS"],
                                        commun_data["ITBIS_RETENIDO"],
                                        commun_data["RETENCION_RENTA"],
                                        invoice_id.id,
                                        AFFECTED_NVOICE_ID and AFFECTED_NVOICE_ID or None,
                                        AFFECTED_NVOICE_ID and True or False))
                purchase_line += 1

            # _logger.info("DGII report {} - - {}".format(count, invoice_id.type))
            count -= 1

        if purchase_report:
            self.create_purchase_lines(purchase_report)

        if sale_report:
            self.create_sales_lines(sale_report)

        self.generate_txt_files()
        # pp(xls_dict)
        self.generate_xls_files(xls_dict)

        if error_list:
            self.post_error_list(error_list)

    def generate_xls_files(self, xls_dict):
        # fill IT-1 excel file
        cwf = os.path.join(os.path.dirname(os.path.abspath(__file__)), "IT-1-2017.xlsx")
        wb = load_workbook(cwf)
        ws1 = wb["IT-1"]  # Get sheet 1 in writeable copy
        xls_dict["it1"].update({"S43": self.positive_balance})
        for k, v in xls_dict["it1"].iteritems():
            ws1[k] = v

        period = self.name.split("/")
        FILENAME = "IT-1-{}-{}.xlsx".format(period[0], period[1])
        wb.save("/tmp/{}".format(FILENAME))
        with open("/tmp/{}".format(FILENAME), "rb") as xls_file:
            self.write({
                'it_filename': FILENAME,
                'it_binary': base64.b64encode(xls_file.read())
            })

        # fill IR-17 excel file
        cwf = os.path.join(os.path.dirname(os.path.abspath(__file__)), "IR-17-2015.xlsx")
        wb = load_workbook(cwf)
        ws1 = wb["IR17"]  # Get sheet 1 in writeable copy
        for k, v in xls_dict["ir17"].iteritems():
            ws1[k] = v

        period = self.name.split("/")
        FILENAME = "IR-17-{}-{}.xlsx".format(period[0], period[1])
        wb.save("/tmp/{}".format(FILENAME))
        with open("/tmp/{}".format(FILENAME), "rb") as xls_file:
            self.write({
                'ir17_filename': FILENAME,
                'ir17_binary': base64.b64encode(xls_file.read())
            })

    def generate_txt_files(self):
        company_fiscal_identificacion = re.sub("[^0-9]", "", self.company_id.vat)
        period = self.name.split("/")
        month = period[0]
        year = period[1]

        sale_path = '/tmp/607{}.txt'.format(company_fiscal_identificacion)
        sale_file = open(sale_path, 'w')

        lines = []

        CANTIDAD_REGISTRO = str(len(self.sale_report)).zfill(12)
        TOTAL_MONTO_FACTURADO_FACTURAS = sum(
            [rec.MONTO_FACTURADO for rec in self.sale_report if rec.NUMERO_COMPROBANTE_MODIFICADO == False])
        TOTAL_MONTO_FACTURADO_NC = sum(
            [rec.MONTO_FACTURADO for rec in self.sale_report if rec.NUMERO_COMPROBANTE_MODIFICADO != False])
        TOTAL_MONTO_FACTURADO = "{:.2f}".format(TOTAL_MONTO_FACTURADO_FACTURAS - TOTAL_MONTO_FACTURADO_NC).zfill(16)

        header = "607"
        header += company_fiscal_identificacion.rjust(11)
        header += str(year)
        header += str(month).zfill(2)
        header += CANTIDAD_REGISTRO
        header += TOTAL_MONTO_FACTURADO
        lines.append(header)

        for sale_line in self.sale_report:
            ln = ""
            ln += sale_line.RNC_CEDULA and sale_line.RNC_CEDULA.rjust(11) or "".rjust(11)
            ln += sale_line.TIPO_IDENTIFICACION
            ln += sale_line.NUMERO_COMPROBANTE_FISCAL.rjust(19)
            ln += sale_line.NUMERO_COMPROBANTE_MODIFICADO or "".rjust(19)
            ln += sale_line.FECHA_COMPROBANTE.replace("-", "")
            ln += "{:.2f}".format(sale_line.ITBIS_FACTURADO).zfill(12)
            ln += "{:.2f}".format(sale_line.MONTO_FACTURADO).zfill(12)
            lines.append(ln)

        for line in lines:
            sale_file.write(line + "\n")

        sale_file.close()
        sale_file = open(sale_path, 'rb')
        sale_binary = base64.b64encode(sale_file.read())
        report_name = 'DGII_607_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year),
                                                    str(month).zfill(2))
        self.write({'sale_binary': sale_binary, 'sale_filename': report_name})

        pruchase_path = '/tmp/606{}.txt'.format(company_fiscal_identificacion)
        purchase_file = open(pruchase_path, 'w')
        lines = []

        CANTIDAD_REGISTRO = "{:.2f}".format(len(self.purchase_report)).zfill(12)
        TOTAL_MONTO_FACTURADO_FACTURAS = sum(
            [rec.MONTO_FACTURADO for rec in self.purchase_report if rec.NUMERO_COMPROBANTE_MODIFICADO == False])
        TOTAL_MONTO_FACTURADO_NC = sum(
            [rec.MONTO_FACTURADO for rec in self.purchase_report if rec.NUMERO_COMPROBANTE_MODIFICADO != False])
        TOTAL_MONTO_FACTURADO = "{:.2f}".format(TOTAL_MONTO_FACTURADO_FACTURAS - TOTAL_MONTO_FACTURADO_NC).zfill(16)

        RETENCION_RENTA = "{:.2f}".format(sum([rec.RETENCION_RENTA for rec in self.purchase_report])).zfill(12)

        header = "606"
        header += company_fiscal_identificacion.rjust(11)
        header += str(year)
        header += str(month).zfill(2)
        header += CANTIDAD_REGISTRO
        header += TOTAL_MONTO_FACTURADO
        header += RETENCION_RENTA
        lines.append(header)

        for line in self.purchase_report:
            ln = ""
            ln += line.RNC_CEDULA.rjust(11)
            ln += line.TIPO_IDENTIFICACION
            ln += line.TIPO_BIENES_SERVICIOS_COMPRADOS
            ln += line.NUMERO_COMPROBANTE_FISCAL and line.NUMERO_COMPROBANTE_FISCAL.rjust(19) or "".rjust(19)
            ln += line.NUMERO_COMPROBANTE_MODIFICADO or "".rjust(19)
            ln += line.FECHA_COMPROBANTE.replace("-", "")
            ln += line.FECHA_PAGO.replace("-", "") if line.FECHA_PAGO else "".rjust(8)
            ln += "{:.2f}".format(line.ITBIS_FACTURADO_TOTAL).zfill(12)
            ln += "{:.2f}".format(abs(line.ITBIS_RETENIDO)).zfill(12)
            ln += "{:.2f}".format(line.MONTO_FACTURADO).zfill(12)
            ln += "{:.2f}".format(line.RETENCION_RENTA).zfill(12)
            lines.append(ln)

        for line in lines:
            purchase_file.write(line + "\n")

        purchase_file.close()
        purchase_file = open(pruchase_path, 'rb')
        purchase_binary = base64.b64encode(purchase_file.read())
        purchase_filename = 'DGII_606_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year), str(month).zfill(2))
        self.write({'purchase_binary': purchase_binary, 'purchase_filename': purchase_filename})

        path = '/tmp/608{}.txt'.format(company_fiscal_identificacion)
        file = open(path, 'w')
        lines = []

        header = "608"
        header += company_fiscal_identificacion.zfill(11)
        header += str(year)
        header += str(month).zfill(2)
        lines.append(header)

        for line in self.cancel_report:
            ln = ""
            ln += line.NUMERO_COMPROBANTE_FISCAL
            ln += line.FECHA_COMPROBANTE and line.FECHA_COMPROBANTE.replace("-", "") or ""
            ln += "{}".format(line.TIPO_ANULACION).zfill(2)
            lines.append(ln)

        for line in lines:
            file.write(line + "\n")

        file.close()
        file = open(path, 'rb')
        report = base64.b64encode(file.read())
        report_name = 'DGII_608_{}_{}{}.TXT'.format(company_fiscal_identificacion, str(year), str(month).zfill(2))
        self.write({'cancel_binary': report, 'cancel_filename': report_name})

    def getTipoComprobante(self, purchase):

        if len(purchase.NUMERO_COMPROBANTE_FISCAL) == 19:
            return purchase.NUMERO_COMPROBANTE_FISCAL[9:-8]
        else:
            return purchase.NUMERO_COMPROBANTE_FISCAL[1:3]





class DgiiReportPurchaseLine(models.Model):
    _name = "dgii.report.purchase.line"

    dgii_report_id = fields.Many2one("dgii.report")
    LINE = fields.Integer("Linea")
    TIPO_BIENES_SERVICIOS_COMPRADOS = fields.Char(u"3 - Tipo Bienes/Servicios", size=2)
    RNC_CEDULA = fields.Char(u"1 - RNC", size=11)
    TIPO_IDENTIFICACION = fields.Char(u"2 - Tipo Identificación", size=1)
    NUMERO_COMPROBANTE_FISCAL = fields.Char("4 - NCF", size=19)
    NUMERO_COMPROBANTE_MODIFICADO = fields.Char(u"5 - NCF Modificado", size=19)
    FECHA_COMPROBANTE = fields.Date(u"6 - Fecha NCF")
    FECHA_PAGO = fields.Date(u"7 - Fecha Pago")
    MONTO_FACTURADO_SERVICIOS = fields.Float(u"8 - Monto Facturado (Servicios)")
    MONTO_FACTURADO_BIENES = fields.Float(u"9 - Monto Facturado (Bienes)")
    MONTO_FACTURADO = fields.Float(u"10 - Monto Facturado (Total)")
    ITBIS_FACTURADO_TOTAL = fields.Float(u"11 - ITBIS Facturado (Total)")
    ITBIS_FACTURADO_BIENES = fields.Float(u"ITBIS Facturado (Bienes)")
    ITBIS_FACTURADO_SERVICIOS = fields.Float(u"ITBIS Facturado (Servicios)")
    ITBIS_RETENIDO = fields.Float(u"12 - ITBIS Retenido")
    ITBIS_SUJETO_PROPORCIONALIDAD = fields.Float(u"13 - ITBIS sujeto a Proporcionalidad (Art. 349)")
    ITBIS_LLEVADO_ALCOSTO = fields.Float(u"14 - ITBIS llevado al Costo")
    ITBIS_POR_ADELANTAR = fields.Float(u"15 - ITBIS por Adelantar")
    ITBIS_PERCIBIDO_COMPRAS = fields.Float(u"16 - ITBIS percibido en compras")
    TIPO_RETENCION_ISR = fields.Char(u"17 - Tipo de Retención en ISR", size=2)
    RETENCION_RENTA = fields.Float(u"18 - Monto Retención Renta")
    ISR_PERCIBIDO_COMPRAS = fields.Float(u"19 - ISR Percibido en compras")
    IMPUESTO_ISC = fields.Float(u"20 - Impuesto Selectivo al Consumo")
    IMPUESTO_OTROS = fields.Float(u"21 - Otros Impuesto/Tasas")
    MONTO_PROPINA_LEGAL = fields.Float(u"22 - Monto Propina Legal")
    FORMA_PAGO = fields.Char(u"23 - Forma de Pago", size=2)

    invoice_id = fields.Many2one("account.invoice", "NCF")
    number = fields.Char(related="invoice_id.number", string=" NCF") #todo validate to remove
    inv_partner = fields.Many2one("res.partner", related="invoice_id.partner_id", string="Proveedor")
    affected_nvoice_id = fields.Many2one("account.invoice", "NCF Modificado")
    nc = fields.Boolean()


class DgiiReportSaleLine(models.Model):
    _name = "dgii.report.sale.line"

    dgii_report_id = fields.Many2one("dgii.report")
    LINE = fields.Integer("Linea")
    RNC_CEDULA = fields.Char(u"RNC", size=11)
    TIPO_IDENTIFICACION = fields.Char(u"Tipo Identificación", size=1)
    NUMERO_COMPROBANTE_FISCAL = fields.Char("NCF", size=19)
    NUMERO_COMPROBANTE_MODIFICADO = fields.Char(u"NCF Modificado", size=19)
    FECHA_COMPROBANTE = fields.Date(u"Fecha NCF")
    ITBIS_FACTURADO = fields.Float(u"ITBIS Facturado")
    MONTO_FACTURADO = fields.Float(u"Monto Facturado")
    MONTO_FACTURADO_EXCENTO = fields.Float(u"Monto Facturado Exento")

    invoice_id = fields.Many2one("account.invoice", "NCF")
    currency_id = fields.Many2one('res.currency', string='Currency', related="invoice_id.currency_id",
                                  required=True, readonly=True, states={'draft': [('readonly', False)]},
                                  track_visibility='always') #todo validate to remove

    number = fields.Char(related="invoice_id.number", string=" NCF") #todo validate to remove
    inv_partner = fields.Many2one("res.partner", related="invoice_id.partner_id", string="Cliente")
    affected_nvoice_id = fields.Many2one("account.invoice", "NCF Modificado")
    nc = fields.Boolean()


class DgiiCancelReportline(models.Model):
    _name = "dgii.cancel.report.line"

    dgii_report_id = fields.Many2one("dgii.report")
    LINE = fields.Integer("Linea")
    NUMERO_COMPROBANTE_FISCAL = fields.Char("NCF", size=19)
    FECHA_COMPROBANTE = fields.Date("Fecha")
    TIPO_ANULACION = fields.Char(u"Tipo de anulación", size=2)
    invoice_id = fields.Many2one("account.invoice", "Factura")


class DgiiExteriorReportline(models.Model):
    _name = "dgii.exterior.report.line"

    dgii_report_id = fields.Many2one("dgii.report")
    LINE = fields.Integer("Linea")
    TIPO_BIENES_SERVICIOS_COMPRADOS = fields.Char("Tipo", size=2)
    FECHA_COMPROBANTE = fields.Date("Fecha")
    FECHA_PAGO = fields.Date("Pagado")
    RETENCION_RENTA = fields.Float(u"Retención Renta")
    MONTO_FACTURADO = fields.Float("Monto Facturado")
    invoice_id = fields.Many2one("account.invoice", "Factura")


class AccountTax(models.Model):
    _inherit = 'account.tax'

    purchase_tax_type = fields.Selection(
        [('itbis', 'ITBIS Pagado (Bienes)'),
         ('itbis_servicios', 'ITBIS Pagado (Servicios)'),
         ('ritbis', 'ITBIS Retenido'),
         ('isr', 'ISR Retenido'),
         ('rext', 'Remesas al Exterior (Ley 253-12)'),
         ('isc', 'Impuesto Selectivo al Consumo (ISC)'),
         ('cdt', 'Contribución Desarrollo Telecomunicaciones (CDT)'),
         ('propina_legal', 'Monto Propina Legal'),
         ('none', 'No Deducible')],
        default="none", string="Tipo de Impuesto en Compra"
    )